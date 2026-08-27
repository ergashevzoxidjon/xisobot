import io
from decimal import Decimal

from flask import (
    Blueprint, render_template, request, redirect, url_for, flash,
    current_app, send_file,
)
from flask_login import login_required, current_user
from sqlalchemy import func

from sqlalchemy.orm import joinedload, selectinload

from extensions import db
from models import (
    Expense, Order, OrderItem, Payment, Client, Material, StockMove, log_action,
    EXPENSE_CATEGORIES, GENERAL_EXPENSE_CATEGORIES, EXPENSE_ORDER_CATEGORY,
    STATUS_CANCELLED, STOCK_OUT, ZERO, PAYMENT_TRANSFER,
)
from permissions import permission_required
from queries import top_debtors, eager_orders, materials_with_stock
from stock import consume, shortage
from utils import (
    ValidationError, parse_money, parse_date, parse_text, parse_choice, parse_int,
    parse_qty, to_money, today_local, month_bounds, money_str, qty_str, QTY_ZERO,
)

finance_bp = Blueprint("finance", __name__, url_prefix="/moliya")

UZ_MONTHS = [
    "Yanvar", "Fevral", "Mart", "Aprel", "May", "Iyun",
    "Iyul", "Avgust", "Sentabr", "Oktabr", "Noyabr", "Dekabr",
]


# ---------- umumiy hisob-kitob funksiyalari ----------

def income_between(start, end):
    """Tushum — AYNAN to'lov sanasi bo'yicha (buyurtma sanasi bo'yicha emas).
    Bekor qilingan buyurtmalar hisobga olinmaydi."""
    total = (
        db.session.query(func.coalesce(func.sum(Payment.amount), 0))
        .join(Order, Payment.order_id == Order.id)
        .filter(
            Payment.paid_on >= start,
            Payment.paid_on < end,
            Order.status != STATUS_CANCELLED,
            Order.is_deleted.is_(False),
        )
        .scalar()
    )
    return to_money(total)


def expenses_between(start, end):
    total = (
        db.session.query(func.coalesce(func.sum(Expense.amount), 0))
        .filter(Expense.date >= start, Expense.date < end)
        .scalar()
    )
    return to_money(total)


def year_rows(year):
    rows = []
    for m in range(1, 13):
        start, end = month_bounds(year, m)
        income = income_between(start, end)
        expense = expenses_between(start, end)
        rows.append({
            "name": UZ_MONTHS[m - 1],
            "month": m,
            "income": income,
            "expense": expense,
            "profit": income - expense,
        })
    return rows


# ---------- xarajatlar ----------

MAX_EXPENSE_ROWS = 30


def parse_expense_order(form):
    """Formadagi buyurtma tanlovini o'qiydi. Bo'sh bo'lsa — umumiy xarajat."""
    raw = (form.get("order_id") or "").strip()
    if not raw:
        return None
    order_id = parse_int(raw, "Buyurtma", min_value=1)
    order = db.session.get(Order, order_id)
    if not order or order.is_deleted:
        raise ValidationError("Buyurtma: tanlangan buyurtma topilmadi.")
    return order


def expense_rows_from_form(form, require_material=False):
    """Formadagi xarajat qatorlarini o'qiydi: mahsulot, soni, narxi.

    Jami avtomatik hisoblanadi (soni × narxi). Qator ombordagi mahsulotga
    tegishli bo'lsa `material` to'ladi — bunday qator yangi xarajat emas,
    ombordan chiqim sifatida yoziladi (puli kirimda allaqachon yozilgan).

    `require_material=True` bo'lsa (buyurtma xarajati) ombordan tanlanmagan
    qator qabul qilinmaydi — avval mahsulot omborga kirim qilinishi kerak.
    """
    material_ids = form.getlist("expense_material")
    descriptions = form.getlist("expense_description")
    quantities = form.getlist("expense_quantity")
    prices = form.getlist("expense_unit_price")

    count = max(len(material_ids), len(descriptions), len(quantities), len(prices))
    if count > MAX_EXPENSE_ROWS:
        raise ValidationError(f"Bir marta {MAX_EXPENSE_ROWS} tadan ko'p qator kiritib bo'lmaydi.")

    def cell(values, index):
        return (values[index] if index < len(values) else "").strip()

    rows = []
    for index in range(count):
        raw_id = cell(material_ids, index)
        description = cell(descriptions, index)
        qty_text = cell(quantities, index)
        price_text = cell(prices, index)

        if not raw_id and not description and not qty_text and not price_text:
            continue

        row = index + 1
        material = None
        if raw_id:
            material = db.session.get(Material, int(raw_id)) if raw_id.isdigit() else None
            if material is None:
                raise ValidationError(f"{row}-qator: mahsulot topilmadi.")
        elif description:
            # foydalanuvchi ro'yxatdan tanlamay, nomini yozgan bo'lishi mumkin
            material = Material.query.filter(Material.name.ilike(description)).first()

        if material is None and require_material:
            raise ValidationError(
                f"{row}-qator: «{description or '—'}» omborda yo'q. "
                f"Avval Ombor → Kirim orqali qabul qiling."
            )

        quantity = parse_qty(qty_text or "1", f"{row}-qator, soni",
                             min_value=Decimal("0.001"))
        unit_price = parse_money(price_text, f"{row}-qator, narxi",
                                 min_value=Decimal("0.01"))

        rows.append({
            "material": material,
            "description": material.name if material else parse_text(
                description, f"{row}-qator, nomi", required=True, max_length=255),
            "quantity": quantity,
            "unit_price": unit_price,
            "amount": to_money(quantity * unit_price),
        })

    if not rows:
        raise ValidationError("Kamida bitta xarajat qatorini to'ldiring.")
    return rows


def stock_shortages(rows):
    """Qatorlardagi mahsulotlar bo'yicha yetishmayotgan miqdorlar ro'yxati.

    Bir mahsulot bir necha qatorda kelsa, miqdorlar qo'shib tekshiriladi.
    Yozuvdan OLDIN chaqiriladi — aks holda yangi chiqimlar qoldiqni buzadi.
    """
    needed = {}
    for data in rows:
        material = data["material"]
        if material is not None:
            needed[material] = needed.get(material, QTY_ZERO) + data["quantity"]

    messages = []
    for material, qty in needed.items():
        missing = shortage(material, qty)
        if missing > QTY_ZERO:
            messages.append(
                f"«{material.name}»: omborda {qty_str(material.quantity)} {material.unit} "
                f"bor, {qty_str(missing)} {material.unit} yetishmadi — qoldiq minusga o'tdi."
            )
    return messages


def expense_category(form, order):
    """Buyurtmaga bog'langan xarajat turkumi avtomatik qo'yiladi.

    Foydalanuvchi turkumni faqat umumiy xarajatlar uchun tanlaydi —
    buyurtma xarajatida bu maydon ko'rinmaydi.
    """
    if order is not None:
        return EXPENSE_ORDER_CATEGORY
    return parse_choice(form.get("category"), "Turkum", GENERAL_EXPENSE_CATEGORIES)


@finance_bp.route("/xarajatlar")
@login_required
@permission_required("expenses.view")
def expenses_list():
    page = request.args.get("page", 1, type=int)
    category = request.args.get("category", "")

    query = Expense.query.options(
        joinedload(Expense.order).joinedload(Order.client)
    )
    if category and category in EXPENSE_CATEGORIES:
        query = query.filter(Expense.category == category)

    pagination = query.order_by(Expense.date.desc(), Expense.id.desc()).paginate(
        page=page, per_page=current_app.config["PER_PAGE"], error_out=False
    )

    total_query = db.session.query(func.coalesce(func.sum(Expense.amount), 0))
    if category and category in EXPENSE_CATEGORIES:
        total_query = total_query.filter(Expense.category == category)
    total_all = to_money(total_query.scalar())

    # Perechisleniye qilingan xaridlar — qaysi tashkilot hisobidan qancha
    # to'langani (ombor kirimidan kelgan xarajatlar bo'yicha, umumiy —
    # filtrga qaramay, xuddi taminotchi qarzi kabi doimiy ko'rinadi).
    transfer_rows = (
        db.session.query(Expense.paid_via, func.coalesce(func.sum(Expense.amount), 0))
        .filter(Expense.payment_method == PAYMENT_TRANSFER)
        .group_by(Expense.paid_via)
        .order_by(func.sum(Expense.amount).desc())
        .all()
    )
    transfer_totals = [{"company": c or "Noma'lum", "total": to_money(t)} for c, t in transfer_rows]
    transfer_total_all = sum((row["total"] for row in transfer_totals), ZERO)

    return render_template(
        "finance/expenses.html",
        expenses=pagination.items,
        pagination=pagination,
        # Filtr tugmalarida "buyurtma" ko'rsatilmaydi — u avtomatik qo'yiladigan
        # ichki turkum, foydalanuvchi tanlamaydi (2026-08-27, foydalanuvchi qarori).
        categories=GENERAL_EXPENSE_CATEGORIES,
        category=category,
        total_all=total_all,
        transfer_totals=transfer_totals,
        transfer_total_all=transfer_total_all,
    )


@finance_bp.route("/xarajatlar/yangi", methods=["GET", "POST"])
@login_required
@permission_required("expenses.create")
def new_expense():
    if request.method == "POST":
        order = None
        try:
            order = parse_expense_order(request.form)
            category = expense_category(request.form, order)
            # Buyurtma xarajatida mahsulot faqat ombordan olinadi
            rows = expense_rows_from_form(request.form, require_material=order is not None)
            exp_date = parse_date(request.form.get("date"), "Sana", required=False) or today_local()
        except ValidationError as e:
            flash(str(e), "danger")
            return _expense_page(selected_order=order, expense=None, form=request.form)

        if exp_date > today_local():
            flash("Xarajat sanasi kelajakda bo'lishi mumkin emas.", "danger")
            return _expense_page(selected_order=order, expense=None, form=request.form)

        # yozuvdan oldin: qoldiq yetadimi
        warnings = stock_shortages(rows)

        total = ZERO
        from_stock = 0
        direct = 0
        for data in rows:
            if data["material"] is not None:
                # ombordan chiqim — puli kirimda yozilgan, yangi xarajat yo'q
                consume(data["material"], data["quantity"], data["unit_price"],
                        order, exp_date)
                from_stock += 1
            else:
                e = Expense(category=category, amount=data["amount"],
                            description=data["description"], date=exp_date,
                            order_id=order.id if order else None,
                            created_by=current_user.id)
                db.session.add(e)
                db.session.flush()
                log_action(current_user, "create", "expense", e.id,
                           f"{category} — {data['amount']}"
                           + (f" ({order.order_number})" if order else ""))
                direct += 1
            total += data["amount"]

        db.session.commit()
        for message in warnings:
            flash(message, "warning")

        parts = []
        if from_stock:
            parts.append(f"{from_stock} ta mahsulot ombordan sarflandi")
        if direct:
            parts.append(f"{direct} ta xarajat yozildi")
        summary = ", ".join(parts) + f" — jami {money_str(total)} so'm."

        if order:
            flash(f"{order.order_number}: {summary}", "success")
            return redirect(url_for("orders.order_detail", order_id=order.id))
        flash(summary[0].upper() + summary[1:], "success")
        if from_stock and not direct:
            return redirect(url_for("stock.list_materials"))
        return redirect(url_for("finance.expenses_list"))

    # Buyurtma sahifasidagi "Xarajat qo'shish" tugmasi o'sha buyurtmani tanlab beradi
    preselected_id = request.args.get("order_id", type=int)
    selected_order = None
    if preselected_id:
        selected_order = (
            Order.query.options(joinedload(Order.client), selectinload(Order.items))
            .get(preselected_id)
        )
    return _expense_page(selected_order=selected_order, expense=None, form=None)


def _expense_page(selected_order, expense, form):
    """Xarajat formasi — ombordagi mahsulotlar ro'yxati bilan."""
    return render_template(
        "finance/expense_form.html",
        categories=GENERAL_EXPENSE_CATEGORIES,
        materials=materials_with_stock(only_active=True),
        selected_order=selected_order, expense=expense, form=form,
    )


@finance_bp.route("/xarajatlar/<int:expense_id>/tahrirlash", methods=["GET", "POST"])
@login_required
@permission_required("expenses.create")
def edit_expense(expense_id):
    e = Expense.query.get_or_404(expense_id)

    if request.method == "POST":
        try:
            order = parse_expense_order(request.form)
            category = expense_category(request.form, order)
            rows = expense_rows_from_form(request.form)
            exp_date = parse_date(request.form.get("date"), "Sana", required=False) or e.date
        except ValidationError as err:
            db.session.rollback()
            flash(str(err), "danger")
            return _expense_page(selected_order=e.order, expense=e, form=request.form)

        # Tahrirlashda bitta yozuv o'zgaradi — birinchi qator hisobga olinadi
        e.order_id = order.id if order else None
        e.category = category
        e.amount = rows[0]["amount"]
        e.description = rows[0]["description"]
        e.date = exp_date

        log_action(current_user, "update", "expense", e.id, f"{e.category} — {e.amount}")
        db.session.commit()
        flash("Xarajat yangilandi.", "success")
        if e.order_id:
            return redirect(url_for("orders.order_detail", order_id=e.order_id))
        return redirect(url_for("finance.expenses_list"))

    return _expense_page(selected_order=e.order, expense=e, form=None)


# ---------- moliyaviy hisobot ----------

@finance_bp.route("/hisobot")
@login_required
@permission_required("reports.view")
def report():
    year = request.args.get("year", today_local().year, type=int)
    if year < 2000 or year > 2100:
        year = today_local().year

    months = year_rows(year)
    total_income = sum((m["income"] for m in months), ZERO)
    total_expense = sum((m["expense"] for m in months), ZERO)
    total_profit = total_income - total_expense

    start, end = date_range_for_year(year)

    return render_template(
        "finance/report.html",
        months=months, year=year,
        total_income=total_income, total_expense=total_expense, total_profit=total_profit,
        **expense_breakdown(start, end),
    )


def expense_breakdown(start, end):
    """Xarajatning uch kesimi: turkum, buyurtma va mahsulot turi.

    Ilgari alohida "Xarajat tahlili" sahifasi bo'lgan — moliyaviy hisobot
    bilan bir xil ishni qilgani uchun (2026-08-26) shu yerga ko'chirildi.
    """
    total = expenses_between(start, end)

    by_category = (
        db.session.query(
            Expense.category,
            func.coalesce(func.sum(Expense.amount), 0),
            func.count(Expense.id),
        )
        .filter(Expense.date >= start, Expense.date < end)
        .group_by(Expense.category)
        .order_by(func.sum(Expense.amount).desc())
        .all()
    )
    category_rows = [
        {"name": name, "amount": to_money(amount), "count": count}
        for name, amount, count in by_category
    ]

    # buyurtmaga bog'langan xarajatlar
    per_order = (
        db.session.query(Expense.order_id, func.coalesce(func.sum(Expense.amount), 0))
        .filter(Expense.date >= start, Expense.date < end, Expense.order_id.isnot(None))
        .group_by(Expense.order_id)
        .all()
    )
    order_expense = {order_id: to_money(amount) for order_id, amount in per_order}
    linked = sum(order_expense.values(), ZERO)
    general = total - linked

    # Ombordan sarflangan mahsulotlar. Ularning puli kirim paytida xarajat
    # sifatida yozilgan, shuning uchun `total` ga QO'SHILMAYDI — faqat
    # buyurtmaning tannarxiga kiradi.
    per_order_stock = (
        db.session.query(
            StockMove.order_id,
            func.coalesce(func.sum(StockMove.quantity * StockMove.unit_price), 0),
        )
        .filter(StockMove.moved_on >= start, StockMove.moved_on < end,
                StockMove.kind == STOCK_OUT, StockMove.order_id.isnot(None))
        .group_by(StockMove.order_id)
        .all()
    )
    stock_used = ZERO
    for order_id, amount in per_order_stock:
        amount = to_money(amount)
        order_expense[order_id] = order_expense.get(order_id, ZERO) + amount
        stock_used += amount

    orders = []
    if order_expense:
        orders = (
            Order.query.options(joinedload(Order.client), selectinload(Order.items))
            .filter(Order.id.in_(list(order_expense)))
            .all()
        )

    order_rows = []
    product_totals = {}
    for o in orders:
        spent = order_expense.get(o.id, ZERO)
        revenue = to_money(o.total_price)
        order_rows.append({
            "order": o, "spent": spent, "revenue": revenue, "profit": revenue - spent,
        })

        # Buyurtma xarajatini mahsulot turlari orasida ulushga qarab taqsimlaymiz:
        # 100 000 so'mlik buyurtmadagi 60 000 so'mlik vizitkaga xarajatning 60% i tegadi.
        items = o.items
        if not items:
            continue
        items_total = sum((to_money(i.total_price) for i in items), ZERO)
        for item in items:
            if items_total > ZERO:
                share = to_money(spent * to_money(item.total_price) / items_total)
            else:
                share = to_money(spent / len(items))
            row = product_totals.setdefault(
                item.order_type, {"spent": ZERO, "revenue": ZERO, "count": 0}
            )
            row["spent"] += share
            row["revenue"] += to_money(item.total_price)
            row["count"] += 1

    order_rows.sort(key=lambda r: r["spent"], reverse=True)

    product_rows = [
        {
            "name": name,
            "spent": values["spent"],
            "revenue": values["revenue"],
            "profit": values["revenue"] - values["spent"],
            "count": values["count"],
        }
        for name, values in product_totals.items()
    ]
    product_rows.sort(key=lambda r: r["spent"], reverse=True)

    return {
        "category_rows": category_rows,
        "linked": linked,
        "general": general,
        "stock_used": stock_used,
        "order_rows": order_rows,
        "product_rows": product_rows,
    }


def date_range_for_year(year):
    from datetime import date
    return date(year, 1, 1), date(year + 1, 1, 1)


# ---------- tahlil ----------

@finance_bp.route("/tahlil")
@login_required
@permission_required("reports.view")
def analytics():
    year = request.args.get("year", today_local().year, type=int)
    start, end = date_range_for_year(year)

    # eng ko'p tushum keltirgan mijozlar
    top_clients = (
        db.session.query(Client.id, Client.name, func.coalesce(func.sum(Payment.amount), 0).label("total"))
        .join(Order, Order.client_id == Client.id)
        .join(Payment, Payment.order_id == Order.id)
        .filter(Payment.paid_on >= start, Payment.paid_on < end,
                Order.status != STATUS_CANCELLED, Order.is_deleted.is_(False))
        .group_by(Client.id, Client.name)
        .order_by(func.sum(Payment.amount).desc())
        .limit(10)
        .all()
    )

    # mahsulot turlari bo'yicha — buyurtma qatorlari kesimida hisoblanadi,
    # shuning uchun bir buyurtmadagi har bir mahsulot alohida ko'rinadi
    top_types = (
        db.session.query(
            OrderItem.order_type,
            func.count(OrderItem.id).label("cnt"),
            func.coalesce(func.sum(OrderItem.total_price), 0).label("total"),
        )
        .join(Order, OrderItem.order_id == Order.id)
        .filter(Order.created_at >= start, Order.created_at < end,
                Order.status != STATUS_CANCELLED, Order.is_deleted.is_(False))
        .group_by(OrderItem.order_type)
        .order_by(func.sum(OrderItem.total_price).desc())
        .limit(10)
        .all()
    )

    total_orders = Order.query.filter(
        Order.created_at >= start, Order.created_at < end, Order.is_deleted.is_(False)
    ).count()
    cancelled = Order.query.filter(
        Order.created_at >= start, Order.created_at < end,
        Order.status == STATUS_CANCELLED, Order.is_deleted.is_(False)
    ).count()
    countable = total_orders - cancelled

    revenue_sum = to_money(
        db.session.query(func.coalesce(func.sum(Order.total_price), 0))
        .filter(Order.created_at >= start, Order.created_at < end,
                Order.status != STATUS_CANCELLED, Order.is_deleted.is_(False))
        .scalar()
    )
    avg_order = to_money(revenue_sum / countable) if countable else ZERO
    cancel_rate = round(cancelled * 100 / total_orders, 1) if total_orders else 0

    # qarzdor mijozlar — bitta agregat so'rov (avval 5201 ta edi)
    debtors = top_debtors(limit=10)

    return render_template(
        "finance/analytics.html",
        year=year,
        top_clients=[{"name": n, "total": to_money(t)} for _, n, t in top_clients],
        top_types=[{"name": t or "-", "count": c, "total": to_money(s)} for t, c, s in top_types],
        total_orders=total_orders, cancelled=cancelled, cancel_rate=cancel_rate,
        avg_order=avg_order, revenue_sum=revenue_sum,
        debtors=debtors,
    )


# ---------- Excel eksport ----------

@finance_bp.route("/hisobot/excel")
@login_required
@permission_required("reports.export")
def export_report():
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill

    year = request.args.get("year", today_local().year, type=int)
    months = year_rows(year)

    wb = Workbook()
    ws = wb.active
    ws.title = f"{year} moliyaviy hisobot"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2563EB")

    ws.append([f"{year}-yil moliyaviy hisoboti"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])

    headers = ["Oy", "Tushum", "Xarajat", "Foyda"]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=3, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for m in months:
        ws.append([m["name"], float(m["income"]), float(m["expense"]), float(m["profit"])])

    total_income = sum((m["income"] for m in months), ZERO)
    total_expense = sum((m["expense"] for m in months), ZERO)
    ws.append([])
    ws.append(["JAMI", float(total_income), float(total_expense), float(total_income - total_expense)])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)

    for col, width in zip("ABCD", (18, 18, 18, 18)):
        ws.column_dimensions[col].width = width
    for row in ws.iter_rows(min_row=4, min_col=2, max_col=4):
        for cell in row:
            cell.number_format = "#,##0.00"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    log_action(current_user, "export", "report", None, f"{year} Excel")
    db.session.commit()
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"moliyaviy_hisobot_{year}.xlsx",
    )


@finance_bp.route("/xarajatlar/excel")
@login_required
@permission_required("reports.export")
def export_expenses():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    expenses = Expense.query.order_by(Expense.date.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Xarajatlar"
    ws.append(["Sana", "Turkum", "Summa", "Izoh", "Kiritgan"])
    for col in range(1, 6):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2563EB")

    for e in expenses:
        ws.append([
            e.date.strftime("%d.%m.%Y"),
            e.category,
            float(to_money(e.amount)),
            e.description or "",
            e.creator.display_name if e.creator else "",
        ])

    for col, width in zip("ABCDE", (14, 16, 16, 40, 20)):
        ws.column_dimensions[col].width = width
    for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):
        for cell in row:
            cell.number_format = "#,##0.00"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="xarajatlar.xlsx",
    )


@finance_bp.route("/buyurtmalar/excel")
@login_required
@permission_required("reports.export")
def export_orders():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    # sana oralig'i bo'yicha cheklash (katta bazada xotirani himoya qiladi)
    try:
        date_from = parse_date(request.args.get("from"), "Boshlanish", required=False)
        date_to = parse_date(request.args.get("to"), "Tugash", required=False)
    except ValidationError as e:
        flash(str(e), "danger")
        return redirect(url_for("orders.list_orders"))

    q = Order.query.filter(Order.is_deleted.is_(False))
    if date_from:
        q = q.filter(Order.created_at >= date_from)
    if date_to:
        q = q.filter(Order.created_at < date_to)

    orders = eager_orders(q.order_by(Order.created_at.desc()).limit(20000)).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Buyurtmalar"
    ws.append(["№", "Sana", "Mijoz", "Turi", "Miqdor", "Summa", "To'langan", "Qolgan", "Holat", "Muddat"])
    for col in range(1, 11):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2563EB")

    for o in orders:
        ws.append([
            o.order_number,
            o.created_at.strftime("%d.%m.%Y") if o.created_at else "",
            o.client.name if o.client else "",
            o.order_type or "",
            o.quantity,
            float(to_money(o.total_price)),
            float(o.paid_amount_calc),
            float(o.remaining),
            o.status,
            o.deadline.strftime("%d.%m.%Y") if o.deadline else "",
        ])

    for col, width in zip("ABCDEFGHIJ", (14, 12, 26, 18, 10, 16, 16, 16, 14, 12)):
        ws.column_dimensions[col].width = width
    for row in ws.iter_rows(min_row=2, min_col=6, max_col=8):
        for cell in row:
            cell.number_format = "#,##0.00"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="buyurtmalar.xlsx",
    )
